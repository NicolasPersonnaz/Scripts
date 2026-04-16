import openpyxl
import paramiko
"""
se fichier note les adresses ip sur lequel il a pu se connecter
 il teste avec A et B il les notes dans excel sur bureau
 
 """
def supprimer_espaces(liste):
    """
    Supprime les espaces avant et après chaque élément d'une liste.

    Args:
        liste (list): La liste à traiter.

    Returns:
        list: La liste modifiée sans les espaces avant et après chaque élément.
    """
    # Utiliser une compréhension de liste avec une condition pour ignorer les éléments vides
    liste_sans_espaces = [element.strip() if element is not None else None for element in liste]
    
    # Filtrer les éléments non-None de la liste
    return [element for element in liste_sans_espaces if element is not None]





def ecrire_liste_dans_excel(liste, nom_fichier):
    """
    Écrit les éléments d'une liste dans un fichier Excel.

    Args:
        liste (list): Liste des éléments à écrire.
        nom_fichier (str): Chemin du fichier Excel à créer ou à écrire.

    Returns:
        None
    """
    # Créer un nouveau classeur Excel
    classeur = openpyxl.Workbook()
    feuille = classeur.active

    # Écrire les éléments de la liste dans la colonne A, un élément par ligne
    for index, element in enumerate(liste, start=1):
        feuille.cell(row=index, column=1, value=element)

    # Enregistrer le fichier Excel
    classeur.save(nom_fichier)

# Exemple d'utilisation :


def tester_connexion(adresses_ip):
    adresses_reussies = []

    with open(r'C:\Users\PERSONNAZN\Desktop\switch_alcatel.xlsx', "a") as fichier:  # Mode ajout
        for ip in adresses_ip:
            try:
                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                
                client.connect(hostname=ip, username='username', password='password', timeout=10)
                
                adresses_reussies.append(ip)
                fichier.write(ip + '\n')
                
                client.close()
            except Exception as e:
                print(f"Erreur lors de la connexion à {ip}: {e}")

    return adresses_reussies

workbook = openpyxl.load_workbook(r'C:\Users\PERSONNAZN\Desktop\test.xlsx')
sheet = workbook.active

adresses_a_tester = []
for i, row in enumerate(sheet.iter_rows(values_only=True)):
   
    ip_address = row[0]
    adresses_a_tester.append(ip_address)
    

adresses_a_tester = supprimer_espaces(adresses_a_tester)



# Tester les connexions SSH avec les adresses IP lues
adresses_reussies = tester_connexion(adresses_a_tester)
print("Connexions réussies : ", adresses_reussies)
ma_liste = adresses_reussies
ecrire_liste_dans_excel(ma_liste, r'C:\Users\PERSONNAZN\Desktop\excel.xlsx')

